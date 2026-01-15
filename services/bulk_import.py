"""
Bulk Import Service for importing contacts from CSV and Excel files.
"""

import csv
import io
import logging
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime

from data.schema import Contact, ImportResult, InputSource
from services.google_sheets import get_sheets_service

logger = logging.getLogger("network_agent.bulk_import")


# Header mapping - maps various column names to Contact fields
HEADER_MAPPINGS = {
    # Name variations
    "name": "full_name",
    "full name": "full_name",
    "contact name": "full_name",
    "full_name": "full_name",
    "fullname": "full_name",
    "first name": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "last name": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",

    # Email variations
    "email": "email",
    "e-mail": "email",
    "email address": "email",
    "mail": "email",
    "e mail": "email",

    # Company variations
    "company": "company",
    "organization": "company",
    "org": "company",
    "firm": "company",
    "company name": "company",
    "organisation": "company",

    # Title variations
    "title": "title",
    "job title": "title",
    "position": "title",
    "role": "title",
    "job_title": "title",
    "jobtitle": "title",

    # Phone variations
    "phone": "phone",
    "mobile": "phone",
    "tel": "phone",
    "telephone": "phone",
    "phone number": "phone",
    "phone_number": "phone",
    "cell": "phone",

    # LinkedIn variations
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "linkedin_url": "linkedin_url",
    "linkedin profile": "linkedin_url",
    "linkedinurl": "linkedin_url",
    "contact_linkedin_url": "linkedin_url",

    # Company LinkedIn
    "company linkedin": "company_linkedin_url",
    "company_linkedin_url": "company_linkedin_url",
    "company linkedin url": "company_linkedin_url",

    # Location variations
    "location": "address",
    "address": "address",
    "city": "address",
    "country": "address",

    # Type variations
    "type": "contact_type",
    "contact type": "contact_type",
    "contact_type": "contact_type",
    "category": "contact_type",
    "classification": "contact_type",

    # Notes
    "notes": "notes",
    "note": "notes",
    "comments": "notes",
    "comment": "notes",

    # Industry
    "industry": "industry",
    "sector": "industry",

    # Website
    "website": "website",
    "web": "website",
    "url": "website",
    "site": "website",

    # Company description
    "company description": "company_description",
    "company_description": "company_description",
    "description": "company_description",
}


class BulkImportService:
    """Service for bulk importing contacts from CSV and Excel files."""

    def __init__(self):
        self.sheets_service = get_sheets_service()

    async def import_file(self, file_bytes: bytes, filename: str) -> ImportResult:
        """
        Main entry point - routes to correct parser based on file extension.

        Args:
            file_bytes: Raw file content
            filename: Original filename (for extension detection)

        Returns:
            ImportResult with counts and any errors
        """
        result = ImportResult()
        filename_lower = filename.lower()

        try:
            if filename_lower.endswith('.csv'):
                rows = self._parse_csv(file_bytes)
            elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
                rows = self._parse_xlsx(file_bytes)
            else:
                result.add_error(f"Unsupported file format: {filename}")
                return result

            if not rows:
                result.add_error("No data found in file")
                return result

            result.total_rows = len(rows)
            logger.info(f"[BULK_IMPORT] Parsed {len(rows)} rows from {filename}")

            # Process each row
            for i, row_data in enumerate(rows):
                try:
                    contact = self._create_contact(row_data)
                    if contact is None:
                        result.skipped += 1
                        continue

                    status, success = self._save_contact(contact)
                    if success:
                        if status == "added":
                            result.successful += 1
                        elif status == "updated":
                            result.updated += 1
                    else:
                        result.failed += 1
                        result.add_error(f"Row {i+2}: Failed to save {contact.name or contact.email}")

                except Exception as e:
                    result.failed += 1
                    result.add_error(f"Row {i+2}: {str(e)}")
                    logger.error(f"[BULK_IMPORT] Error processing row {i+2}: {e}")

            logger.info(f"[BULK_IMPORT] Complete: {result.summary()}")

        except Exception as e:
            result.add_error(f"Failed to parse file: {str(e)}")
            logger.error(f"[BULK_IMPORT] Parse error: {e}")

        return result

    def _parse_csv(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """
        Parse CSV file into list of contact dictionaries.

        Args:
            file_bytes: Raw CSV content

        Returns:
            List of dictionaries with mapped field names
        """
        try:
            content = file_bytes.decode('utf-8')
        except UnicodeDecodeError:
            # Try with latin-1 encoding
            content = file_bytes.decode('latin-1')

        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        if not rows:
            return []

        # First row is headers
        raw_headers = rows[0]
        header_map = self._detect_headers(raw_headers)

        logger.info(f"[BULK_IMPORT] CSV headers detected: {header_map}")

        # Parse remaining rows
        contacts = []
        for row in rows[1:]:
            if not any(cell.strip() for cell in row):
                continue  # Skip empty rows

            contact_data = {}
            for i, cell in enumerate(row):
                if i < len(raw_headers):
                    header = raw_headers[i].lower().strip()
                    field_name = header_map.get(header)
                    if field_name and cell.strip():
                        contact_data[field_name] = cell.strip()

            if contact_data:
                contacts.append(contact_data)

        return contacts

    def _parse_xlsx(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """
        Parse Excel file into list of contact dictionaries.

        Args:
            file_bytes: Raw Excel content

        Returns:
            List of dictionaries with mapped field names
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("[BULK_IMPORT] openpyxl not installed")
            raise ImportError("openpyxl is required for Excel files. Run: pip install openpyxl")

        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        # First row is headers
        raw_headers = [str(h).strip() if h else "" for h in rows[0]]
        header_map = self._detect_headers(raw_headers)

        logger.info(f"[BULK_IMPORT] Excel headers detected: {header_map}")

        # Parse remaining rows
        contacts = []
        for row in rows[1:]:
            if not any(cell for cell in row if cell is not None):
                continue  # Skip empty rows

            contact_data = {}
            for i, cell in enumerate(row):
                if i < len(raw_headers):
                    header = raw_headers[i].lower().strip()
                    field_name = header_map.get(header)
                    if field_name and cell is not None:
                        value = str(cell).strip()
                        if value:
                            contact_data[field_name] = value

            if contact_data:
                contacts.append(contact_data)

        workbook.close()
        return contacts

    def _detect_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Map file headers to Contact field names.

        Args:
            headers: List of column headers from file

        Returns:
            Dictionary mapping lowercase header -> Contact field name
        """
        header_map = {}
        for header in headers:
            if not header:
                continue
            header_lower = header.lower().strip()
            if header_lower in HEADER_MAPPINGS:
                header_map[header_lower] = HEADER_MAPPINGS[header_lower]
            else:
                # Check if it's already a valid field name
                valid_fields = [
                    "full_name", "first_name", "last_name", "email", "phone",
                    "company", "title", "linkedin_url", "address", "notes",
                    "industry", "website", "contact_type", "company_description"
                ]
                if header_lower in valid_fields:
                    header_map[header_lower] = header_lower

        return header_map

    def _create_contact(self, row_data: Dict[str, Any]) -> Optional[Contact]:
        """
        Create a Contact object from parsed row data.

        Args:
            row_data: Dictionary with field names and values

        Returns:
            Contact object or None if invalid
        """
        # Must have at least name OR email
        full_name = row_data.get("full_name")
        first_name = row_data.get("first_name")
        last_name = row_data.get("last_name")
        email = row_data.get("email")

        # Build full name if not provided
        if not full_name and (first_name or last_name):
            full_name = f"{first_name or ''} {last_name or ''}".strip()

        # Must have at least one identifier
        if not full_name and not email:
            return None

        # Create contact
        contact = Contact(
            first_name=first_name,
            last_name=last_name,
            full_name=full_name,
            email=email,
            phone=row_data.get("phone"),
            company=row_data.get("company"),
            title=row_data.get("title"),
            linkedin_url=row_data.get("linkedin_url"),
            address=row_data.get("address"),
            notes=row_data.get("notes"),
            industry=row_data.get("industry"),
            website=row_data.get("website"),
            contact_type=row_data.get("contact_type"),
            company_description=row_data.get("company_description"),
            source=InputSource.BULK.value,
            imported_date=datetime.now().strftime("%Y-%m-%d"),
        )

        return contact

    def _save_contact(self, contact: Contact) -> Tuple[str, bool]:
        """
        Save or update a contact.

        Args:
            contact: Contact object to save

        Returns:
            Tuple of (status, success) where status is "added", "updated", or "skipped"
        """
        # Check for existing by email first
        if contact.email:
            existing = self.sheets_service.find_contact_by_email(contact.email)
            if existing:
                # Update existing contact with non-empty fields
                updates = self._get_non_empty_fields(contact)
                if updates and self.sheets_service.update_contact(existing.name, updates):
                    logger.info(f"[BULK_IMPORT] Updated existing contact: {existing.name}")
                    return "updated", True
                return "skipped", False

        # Check for existing by name
        name = contact.full_name or contact.name
        if name:
            existing = self.sheets_service.get_contact_by_name(name)
            if existing:
                updates = self._get_non_empty_fields(contact)
                if updates and self.sheets_service.update_contact(existing.name, updates):
                    logger.info(f"[BULK_IMPORT] Updated existing contact: {existing.name}")
                    return "updated", True
                return "skipped", False

        # Add as new contact
        if self.sheets_service.add_contact(contact):
            logger.info(f"[BULK_IMPORT] Added new contact: {contact.name}")
            return "added", True

        return "skipped", False

    def _get_non_empty_fields(self, contact: Contact) -> Dict[str, Any]:
        """
        Get dictionary of non-empty fields from a contact.

        Args:
            contact: Contact object

        Returns:
            Dictionary of field names to non-empty values
        """
        fields_to_check = [
            "first_name", "last_name", "full_name", "email", "phone",
            "company", "title", "linkedin_url", "address", "notes",
            "industry", "website", "contact_type", "company_description"
        ]

        updates = {}
        for field in fields_to_check:
            value = getattr(contact, field, None)
            if value:
                updates[field] = value

        return updates


# Global service instance
_bulk_import_service: Optional[BulkImportService] = None


def get_bulk_import_service() -> BulkImportService:
    """Get or create the bulk import service instance."""
    global _bulk_import_service
    if _bulk_import_service is None:
        _bulk_import_service = BulkImportService()
    return _bulk_import_service
